import os
import pandas as pd
from datetime import datetime, timedelta
import re
from abc import ABC, abstractmethod
from copy import copy, deepcopy
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import re
from typing import List, Dict, Any
from .csv_handler import CSVHandler
from .excel_handler import ExcelHandler
from .importer_interface import TransactionImporter


class BaseImporter(TransactionImporter):
    """Base class for importing transactions from bank CSV files to Excel."""
    
    def __init__(self):
        self.csv_file = None
        self.excel_file = None
        
        # Default column mappings for common fields
        self.column_mappings = {
            'Date': 'Date',
            'Description': 'Description',
            'Amount': 'Amount',
            'Balance': 'Balance',
            'Account Number': 'Account Number',
            'Transaction Type': 'Transaction Type'
        }
        
        # Default values for required fields
        self.default_values = {
            'Account Number': '',
            'Transaction Type': '',
            'Balance': 0.0
        }
    
    @abstractmethod
    def get_expected_columns(self):
        """Return the expected columns for this bank's CSV file."""
        pass
    
    @abstractmethod
    def get_institution_name(self):
        """Return the institution name for this bank."""
        pass
    
    @abstractmethod
    def get_account_name(self):
        """Return the default account name for this bank."""
        pass
    
    @abstractmethod
    def parse_transaction_amount(self, amount_str, transaction_type=None):
        """Parse transaction amount and determine if it's a debit or credit."""
        pass

    def set_column_mapping(self, source_column: str, target_column: str):
        """Set a custom column mapping for this bank."""
        self.column_mappings[target_column] = source_column
    
    def set_default_value(self, column: str, value: Any):
        """Set a default value for a column if it's missing in the CSV."""
        self.default_values[column] = value
    
    def get_column_value(self, row: Dict[str, Any], column_name: str) -> str:
        """Get a column value from the row, using mapping and default values if needed."""
        # Get the mapped column name
        mapped_column = self.column_mappings.get(column_name)
        
        # If we have a mapping and the column exists in the row, use it
        if mapped_column and mapped_column in row:
            return row[mapped_column]
        
        # If we have a default value, use it
        if column_name in self.default_values:
            return self.default_values[column_name]
        
        # If no mapping or default, return empty string
        return ''
    
    def parse_transaction_date(self, date_str):
        """Parse transaction date from string."""
        try:
            # Try multiple date formats
            date_formats = ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%d/%m/%Y']
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            
            # Last resort - try pandas
            date = pd.to_datetime(date_str, errors='coerce').date()
            if pd.isna(date):
                return datetime.now().date()
            return date
                
        except Exception as date_error:
            print(f"Date parsing error: {date_error}")
            print(f"Raw date value: '{date_str}'")
            return datetime.now().date()

    def validate_files(self, csv_file, excel_file):
        """Validate that input files exist and are accessible."""
        if not os.path.exists(csv_file):
            raise FileNotFoundError(f"CSV file not found: {csv_file}")
        
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"Excel file not found: {excel_file}")
        
        if not csv_file.lower().endswith('.csv'):
            raise ValueError("First file must be a CSV file")
        
        if not excel_file.lower().endswith(('.xlsx', '.xls')):
            raise ValueError("Second file must be an Excel file")
    
    def read_csv(self, csv_file):
        """Read and parse CSV file with multiple encoding attempts."""
        try:
            # Try different encodings in case of special characters
            encodings = ['utf-8', 'latin-1', 'cp1252']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(csv_file, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise Exception("Could not read CSV file with any supported encoding")
            
            # Strip whitespace from column names
            df.columns = df.columns.str.strip()
            
            # Remove any completely empty rows
            df = df.dropna(how='all')
            
            return df
            
        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}") from e
    
    def get_week_start(self, date):
        """Get the first day of the week (Sunday) for a given date."""
        # Get the weekday (0=Monday, 6=Sunday in Python)
        days_since_sunday = (date.weekday() + 1) % 7
        week_start = date - timedelta(days=days_since_sunday)
        return week_start
    
    def format_date_mdy(self, date_obj):
        """Format date as M/D/YYYY (without leading zeros) - cross-platform compatible."""
        if isinstance(date_obj, str):
            return date_obj
        
        # Use manual formatting to avoid platform-specific strftime issues
        month = date_obj.month
        day = date_obj.day
        year = date_obj.year
        return f"{month}/{day}/{year}"
    
    def clean_description(self, description):
        """Clean and format transaction description to be more human readable."""
        if not description or pd.isna(description):
            return ''
        
        description = str(description).strip()
        
        # Remove common prefixes
        prefixes_to_remove = [
            "Digital Card Purchase - ",
            "Debit Card Purchase - ",
            "Credit Card Purchase - ",
            "Card Purchase - ",
            "Purchase - "
        ]
        
        cleaned = description
        for prefix in prefixes_to_remove:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix):]
                break
        
        # Remove common payment processor prefixes (more comprehensive)
        processor_prefixes = [
            "TST* ",    # Toast (with asterisk)
            "TST ",     # Toast (without asterisk)
            "SQ* ",     # Square (with asterisk)
            "SQ ",      # Square (without asterisk)
            "SP* ",     # Stripe (with asterisk)
            "SP ",      # Stripe (without asterisk)
            "PP* ",     # PayPal (with asterisk)
            "PP ",      # PayPal (without asterisk)
            "PAYPAL* ", # PayPal variant (with asterisk)
            "PAYPAL ",  # PayPal variant (without asterisk)
            "AMZN* ",   # Amazon (with asterisk)
            "AMZN ",    # Amazon (without asterisk)
            "UBER* ",   # Uber (with asterisk)
            "UBER ",    # Uber (without asterisk)
            "LYFT* ",   # Lyft (with asterisk)
            "LYFT ",    # Lyft (without asterisk)
        ]
        
        for prefix in processor_prefixes:
            if cleaned.upper().startswith(prefix.upper()):
                cleaned = cleaned[len(prefix):]
                break
        
        # Remove store numbers and common patterns
        # Remove patterns like #243, #1760, etc. (store numbers with #)
        cleaned = re.sub(r'\s*#\d+\s*', ' ', cleaned)
        # Remove standalone numbers that look like store IDs (4-6 digits)
        cleaned = re.sub(r'\s+\d{4,6}\s+', ' ', cleaned)
        
        # Clean up extra spaces
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        
        # Handle some final cleanup
        cleaned = cleaned.replace(',  ', ', ')  # Fix spacing after commas
        
        # Convert to title case (first letter of each word capitalized, rest lowercase)
        cleaned = ' '.join(word.capitalize() for word in cleaned.split())
        
        return cleaned

    def check_for_duplicates(self, existing_df, new_transactions):
        """Check for potential duplicate transactions and return filtered list."""
        if existing_df.empty or not new_transactions:
            return new_transactions
        
        # Convert new transactions to DataFrame for easier comparison
        new_df = pd.DataFrame(new_transactions)
        
        # Create a comparison key using date, amount, and first 20 chars of description
        def create_comparison_key(row):
            date_str = str(row.get('Date', ''))
            amount_str = str(row.get('Amount', ''))
            desc_str = str(row.get('Description', ''))[:20].strip()
            return f"{date_str}|{amount_str}|{desc_str}"
        
        # Create comparison keys for existing transactions
        existing_keys = set()
        for _, row in existing_df.iterrows():
            key = create_comparison_key(row)
            existing_keys.add(key)
        
        # Filter out potential duplicates
        filtered_transactions = []
        duplicate_count = 0
        
        for transaction in new_transactions:
            key = create_comparison_key(transaction)
            if key not in existing_keys:
                filtered_transactions.append(transaction)
                existing_keys.add(key)  # Add to existing keys to prevent duplicates within new transactions
            else:
                duplicate_count += 1
        
        if duplicate_count > 0:
            print(f"Filtered out {duplicate_count} potential duplicate transactions")
        
        return filtered_transactions
    
    def update_excel_file(self, excel_file, transactions):
        """Update the Tiller Excel file with new transactions while preserving formatting."""
        try:
            # Load the workbook and worksheet
            wb = load_workbook(excel_file)
            
            if 'Transactions' not in wb.sheetnames:
                raise ValueError("'Transactions' worksheet not found in Excel file")
            
            ws = wb['Transactions']
            
            # Read the existing data to get column structure and check for duplicates
            existing_df = pd.read_excel(excel_file, sheet_name='Transactions')
            existing_columns = list(existing_df.columns)
            
            # Filter out potential duplicates
            filtered_transactions = self.check_for_duplicates(existing_df, transactions)
            
            if not filtered_transactions:
                print("No new transactions to import (all appear to be duplicates)")
                return 0
            
            # Find the header row (first row with data)
            header_row = 1
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=20):
                if any(cell.value for cell in row):
                    header_row = row[0].row
                    break
            
            print(f"Header row found at: {header_row}")
            
            # Convert new transactions to DataFrame with proper column order
            new_df = pd.DataFrame(filtered_transactions)
            new_df = new_df.reindex(columns=existing_columns, fill_value='')
            
            # Sort new transactions by date (newest first)
            if 'Date' in new_df.columns and not new_df.empty:
                new_df = new_df.sort_values('Date', ascending=False)
            
            # Get the date column index
            date_col_idx = existing_columns.index('Date') + 1  # +1 because Excel is 1-based
            
            # Store the current max row BEFORE making any changes
            original_max_row = ws.max_row
            first_data_row = header_row + 1
            
            # Capture formatting information BEFORE making changes
            print("Capturing existing formatting...")
            
            # 1. Capture data validations and their ranges
            original_data_validations = []
            for dv in ws.data_validations.dataValidation:
                # Store the validation rule and ALL its current ranges
                dv_info = {
                    'type': dv.type,
                    'formula1': dv.formula1,
                    'formula2': dv.formula2,
                    'allow_blank': dv.allow_blank,
                    'showDropDown': dv.showDropDown,
                    'showInputMessage': dv.showInputMessage,
                    'showErrorMessage': dv.showErrorMessage,
                    'errorTitle': dv.errorTitle,
                    'error': dv.error,
                    'promptTitle': dv.promptTitle,
                    'prompt': dv.prompt,
                    'ranges': []
                }
                
                # Collect all ranges for this validation
                for cell_range in dv.ranges:
                    dv_info['ranges'].append(str(cell_range))
                
                original_data_validations.append(dv_info)
                print(f"Found data validation: {dv.type} for ranges: {dv_info['ranges']}")
            
            # 2. Capture conditional formatting rules and their ranges
            original_conditional_formatting = []
            for cf_range in ws.conditional_formatting:
                rules_list = ws.conditional_formatting[cf_range]
                for rule in rules_list:
                    cf_info = {
                        'rule': deepcopy(rule),
                        'range': str(cf_range) if hasattr(cf_range, 'coord') else cf_range
                    }
                    original_conditional_formatting.append(cf_info)
                    print(f"Found conditional formatting rule for range: {cf_info['range']}")
            
            # 3. Capture cell styles from the first data row (template)
            template_styles = {}
            if first_data_row <= original_max_row:
                for col in range(1, len(existing_columns) + 1):
                    cell = ws.cell(row=first_data_row, column=col)
                    if cell:
                        template_styles[col] = {
                            'font': copy(cell.font),
                            'fill': copy(cell.fill),
                            'border': copy(cell.border),
                            'alignment': copy(cell.alignment),
                            'number_format': cell.number_format,
                            'protection': copy(cell.protection)
                        }
            
            # Find empty rows in the table
            empty_rows = []
            for row in range(header_row + 1, ws.max_row + 1):
                if all(ws.cell(row=row, column=col).value is None for col in range(1, len(existing_columns) + 1)):
                    empty_rows.append(row)
            
            # Track which rows we actually insert (for updating ranges later)
            inserted_rows = []
            
            # For each new transaction, find its correct position based on date
            for _, row in new_df.iterrows():
                # Get the transaction date
                trans_date = row['Date']
                if isinstance(trans_date, datetime):
                    trans_date = trans_date.date()
                
                # Find the correct insertion point
                current_row = header_row + 1
                while current_row <= ws.max_row:
                    cell_value = ws.cell(row=current_row, column=date_col_idx).value
                    if cell_value is None:
                        break
                    
                    # Convert cell value to date if it's a string or datetime
                    if isinstance(cell_value, str):
                        try:
                            cell_date = pd.to_datetime(cell_value).date()
                        except:
                            break
                    elif isinstance(cell_value, datetime):
                        cell_date = cell_value.date()
                    else:
                        cell_date = cell_value
                    
                    # If transaction date is newer than cell date, insert here
                    if trans_date > cell_date:
                        break
                    
                    current_row += 1
                
                # Try to reuse an empty row if available
                target_row = None
                for empty_row in empty_rows:
                    if empty_row <= current_row:
                        target_row = empty_row
                        empty_rows.remove(empty_row)
                        break
                
                if target_row is None:
                    # Insert a new row
                    ws.insert_rows(current_row)
                    target_row = current_row
                    inserted_rows.append(target_row)
                    print(f"Inserted new row at position {target_row}")
                
                # Add the transaction data
                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=target_row, column=col_idx)
                    
                    # Apply template style if available
                    if col_idx in template_styles:
                        style = template_styles[col_idx]
                        cell.font = style['font']
                        cell.fill = style['fill'] 
                        cell.border = style['border']
                        cell.alignment = style['alignment']
                        cell.number_format = style['number_format']
                        cell.protection = style['protection']
                    
                    # Handle different data types
                    if pd.isna(value):
                        cell.value = None
                    elif isinstance(value, (pd.Timestamp, datetime)):
                        cell.value = value.date() if hasattr(value, 'date') else value
                    else:
                        cell.value = value
            
            # Now update data validation ranges if we inserted rows
            if inserted_rows and original_data_validations:
                print(f"Updating data validation for {len(inserted_rows)} inserted rows...")
                
                # Clear existing data validations
                ws.data_validations = type(ws.data_validations)()
                
                # Re-create each data validation with expanded ranges
                for dv_info in original_data_validations:
                    new_dv = DataValidation(
                        type=dv_info['type'],
                        formula1=dv_info['formula1'],
                        formula2=dv_info['formula2'],
                        allow_blank=dv_info['allow_blank'],
                        showDropDown=dv_info['showDropDown'],
                        showInputMessage=dv_info['showInputMessage'],
                        showErrorMessage=dv_info['showErrorMessage'],
                        errorTitle=dv_info['errorTitle'],
                        error=dv_info['error'],
                        promptTitle=dv_info['promptTitle'],
                        prompt=dv_info['prompt']
                    )
                    
                    # Process each range and expand if necessary
                    for range_str in dv_info['ranges']:
                        expanded_ranges = self._expand_range_for_new_rows(
                            range_str, inserted_rows, original_max_row, header_row
                        )
                        for expanded_range in expanded_ranges:
                            new_dv.add(expanded_range)
                            print(f"Added validation range: {expanded_range}")
                    
                    ws.add_data_validation(new_dv)
            
            # Update conditional formatting ranges if we inserted rows
            if inserted_rows and original_conditional_formatting:
                print(f"Updating conditional formatting for {len(inserted_rows)} inserted rows...")
                
                # Clear existing conditional formatting
                ws.conditional_formatting = type(ws.conditional_formatting)() 
                
                # Re-create each conditional formatting rule with expanded ranges
                for cf_info in original_conditional_formatting:
                    expanded_ranges = self._expand_range_for_new_rows(
                        cf_info['range'], inserted_rows, original_max_row, header_row
                    )
                    for expanded_range in expanded_ranges:
                        ws.conditional_formatting.add(expanded_range, cf_info['rule'])
                        print(f"Added conditional formatting range: {expanded_range}")
            
            # Find and update table if it exists
            if ws.tables:
                table_name = list(ws.tables.keys())[0]
                table = ws.tables[table_name]
                print(f"Found table: {table_name}")
                
                # Calculate new table range
                max_row = ws.max_row
                max_col = len(existing_columns)
                max_col_letter = get_column_letter(max_col)
                
                # Update table range
                new_range = f"A{header_row}:{max_col_letter}{max_row}"
                table.ref = new_range
                print(f"Extended table range to: {new_range}")
            
            # Save the workbook
            wb.save(excel_file)
            wb.close()
            
            print(f"Successfully added {len(filtered_transactions)} transactions")
            return len(filtered_transactions)
            
        except Exception as e:
            raise Exception(f"Error updating Excel file: {str(e)}") from e

    def _expand_range_for_new_rows(self, range_str, inserted_rows, original_max_row, header_row):
        """Expand a range to include newly inserted rows."""
        from openpyxl.utils import range_boundaries, get_column_letter
        
        try:
            # Clean up the range string if it's a ConditionalFormatting object
            if isinstance(range_str, str) and range_str.startswith('<ConditionalFormatting'):
                # Extract the actual range from the string (e.g., "D2" from "<ConditionalFormatting D2>")
                range_str = range_str.split()[-1].rstrip('>')
            
            # Parse the original range
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
            
            expanded_ranges = []
            
            # Add the original range (updated if it extended to the bottom)
            if max_row >= original_max_row:
                # This range went to the bottom, so extend it
                new_max_row = max_row + len(inserted_rows)
                min_col_letter = get_column_letter(min_col)
                max_col_letter = get_column_letter(max_col)
                extended_range = f"{min_col_letter}{min_row}:{max_col_letter}{new_max_row}"
                expanded_ranges.append(extended_range)
            else:
                # Keep original range as-is
                expanded_ranges.append(range_str)
            
            # Add ranges for each inserted row (if the original range was row-specific)
            if min_row > header_row and max_row < original_max_row:
                # This looks like it was applied to specific data rows
                min_col_letter = get_column_letter(min_col)
                max_col_letter = get_column_letter(max_col)
                
                for inserted_row in inserted_rows:
                    if min_col == max_col:
                        # Single column
                        new_range = f"{min_col_letter}{inserted_row}"
                    else:
                        # Multiple columns
                        new_range = f"{min_col_letter}{inserted_row}:{max_col_letter}{inserted_row}"
                    expanded_ranges.append(new_range)
            
            return expanded_ranges
            
        except Exception as e:
            print(f"Warning: Could not expand range {range_str}: {e}")
            return [range_str]

    def read_bank_csv(self, csv_file):
        """Read and parse bank CSV file."""
        df = self.read_csv(csv_file)
        
        # Check if all expected columns are present
        expected_columns = self.get_expected_columns()
        missing_columns = set(expected_columns) - set(df.columns)
        if missing_columns:
            # Try to find similar column names
            available_cols = list(df.columns)
            suggestion_msg = f"\nAvailable columns in your CSV: {available_cols}"
            error_msg = f"Your CSV file is missing some required columns: {missing_columns}.{suggestion_msg}\n\n"
            error_msg += "This usually means either:\n"
            error_msg += "1. The CSV file is from a different bank than selected\n"
            error_msg += "2. The CSV file format has changed\n"
            error_msg += "3. The CSV file was exported incorrectly\n\n"
            error_msg += "Please check that you selected the correct bank and that the CSV file is properly exported."
            raise ValueError(error_msg)
        
        return df

    def transform_transactions(self, df: pd.DataFrame, existing_columns: List[str]) -> List[Dict[str, Any]]:
        """Transform bank transactions to Tiller format based on existing columns."""
        transformed_transactions = []
        
        for index, row in df.iterrows():
            try:
                # Parse transaction date
                trans_date = self.parse_transaction_date(str(self.get_column_value(row, 'Date')).strip())
                
                # Parse transaction amount
                amount = self.parse_transaction_amount(
                    str(self.get_column_value(row, 'Amount')),
                    str(self.get_column_value(row, 'Transaction Type')).lower().strip()
                )
                
                # Calculate date-related fields
                year_start = datetime(trans_date.year, 1, 1)
                month_start = datetime(trans_date.year, trans_date.month, 1)
                week_start = self.get_week_start(trans_date)
                
                # Create transaction mapping
                transaction = {
                    'Date': trans_date,
                    'Description': self.clean_description(self.get_column_value(row, 'Description')),
                    'Category': '',  # Will be empty for user to categorize
                    'Amount': amount,
                    'Account': self.get_account_name(),
                    'Account #': str(self.get_column_value(row, 'Account Number')),
                    'Institution': self.get_institution_name(),
                    'Year': self.format_date_mdy(year_start),
                    'Month': self.format_date_mdy(month_start),
                    'Week': self.format_date_mdy(week_start),
                    'Check Number': '',  # Not applicable for most transactions
                    'Full Description': str(self.get_column_value(row, 'Description')),
                    'Date Added': self.format_date_mdy(datetime.now())
                }
                
                # Filter to only include columns that exist in the Excel file
                filtered_transaction = {
                    col: transaction.get(col, '')
                    for col in existing_columns
                }
                
                transformed_transactions.append(filtered_transaction)
                
            except Exception as row_error:
                print(f"Error processing row {index}: {row_error}")
                print(f"Row data: {dict(row)}")
                print("Skipping this transaction and continuing...")
                continue
        
        return transformed_transactions

    def import_transactions(self, csv_file: str, excel_file: str) -> tuple[bool, str]:
        """Main import function."""
        try:
            # Initialize handlers
            csv_handler = CSVHandler(csv_file)
            excel_handler = ExcelHandler(excel_file)
            
            # Validate files
            csv_handler.validate_file()
            
            # Read and validate CSV
            print(f"Reading CSV file: {csv_file}")
            df = csv_handler.read_csv()
            csv_handler.validate_columns(self.get_expected_columns())
            print(f"Found {len(df)} transactions in CSV")
            
            # Load Excel file
            excel_handler.load_workbook()
            existing_columns = excel_handler.existing_columns
            print(f"Excel file has {len(existing_columns)} columns: {existing_columns}")
            
            # Transform transactions
            print("Transforming transactions...")
            transactions = self.transform_transactions(df, existing_columns)
            
            # Update Excel file
            print(f"Updating Excel file: {excel_file}")            
            excel_handler.update_transactions(transactions)
            excel_handler.save()
            
            count = len(transactions)
            if count > 0:
                print(f"Successfully imported {count} new transactions!")
                return True, f"Successfully imported {count} new transactions!"
            else:
                print("No new transactions to import (all appear to be duplicates)")
                return True, "No new transactions to import (all appear to be duplicates)"
            
        except Exception as e:
            import traceback
            # Print full traceback to console
            print(f"Error: {str(e)}\nTraceback:\n{traceback.format_exc()}")
            # Return only the error message without traceback
            return False, f"Error: {str(e)}" 