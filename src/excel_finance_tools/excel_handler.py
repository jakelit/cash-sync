import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict, Any, Optional
from .duplicate_checker import DuplicateChecker
from datetime import datetime
from copy import copy
import traceback
from .logger import logger

class ExcelHandler:
    """
    Excel handler that properly works with Excel Tables.
    """
    
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.wb = None
        self.ws = None
        self.table = None
        self.existing_df = None
        self.existing_columns = None
        self.column_mapping = {}  # Maps transaction columns to Excel column indices
        self.header_row = None
    
    def load_workbook(self):
        """Load the Excel workbook and locate the Transactions table."""
        logger.info(f"Loading workbook: {self.excel_file}")
        self.wb = load_workbook(self.excel_file)
        
        # Find the Transactions table across all worksheets
        transactions_table = None
        table_worksheet = None
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if ws.tables:
                logger.debug(f"Found {len(ws.tables)} table(s) in worksheet '{sheet_name}': {list(ws.tables.keys())}")
                for table_name in ws.tables:
                    if table_name.lower() == 'transactions':
                        transactions_table = ws.tables[table_name]
                        table_worksheet = ws
                        logger.info(f"Found 'Transactions' table in worksheet '{sheet_name}'")
                        break
                if transactions_table:
                    break
        
        if not transactions_table:
            self._raise_table_missing_error()
        
        self.table = transactions_table
        self.ws = table_worksheet
        
        logger.debug(f"Table range: {self.table.ref}")
        self._load_existing_data()
        self._build_column_mapping()
    
    def _raise_table_missing_error(self):
        """Raise a clear error with instructions for creating the Transactions table."""
        error_message = """
❌ ERROR: No 'Transactions' table found in the Excel file.

🔧 HOW TO FIX THIS:

1. Open your Excel file
2. Select your transaction data (including headers)
3. Go to Insert → Table (or press Ctrl+T)
4. Make sure "My table has headers" is checked
5. Click OK
6. Right-click the table and select "Table Design"
7. In the "Table Name" box, change the name to: Transactions
8. Save your file

💡 WHY THIS IS REQUIRED:
Excel Tables automatically preserve formatting, data validation, and conditional 
formatting when new rows are added. This eliminates the need for complex 
formatting management in the code.

📋 TABLE REQUIREMENTS:
- Table name must be exactly: Transactions
- Must include column headers
- Should contain your existing transaction data
        """
        raise ValueError(error_message.strip())
    
    def _load_existing_data(self):
        """Load existing data from the Transactions table."""
        # Get table range
        min_col, min_row, max_col, max_row = range_boundaries(self.table.ref)
        
        # Load data from the specific table range
        self.existing_df = pd.read_excel(
            self.excel_file, 
            sheet_name=self.ws.title,
            usecols=f"{get_column_letter(min_col)}:{get_column_letter(max_col)}",
            skiprows=min_row - 1,  # Skip to header row
            nrows=max_row - min_row,  # Read only table rows
            index_col=None  # Ensure we get a default 0-based index
        )
        
        # Store the header row number for later calculations
        self.header_row = min_row
        
        self.existing_columns = list(self.existing_df.columns)
        logger.info(f"Loaded {len(self.existing_df)} existing transactions")
        logger.debug(f"Table columns: {self.existing_columns}")
    
    def _build_column_mapping(self):
        """Build mapping between transaction columns and Excel column indices."""
        min_col, min_row, max_col, max_row = range_boundaries(self.table.ref)
        
        # Create case-insensitive mapping
        self.column_mapping = {}
        for i, col_name in enumerate(self.existing_columns):
            excel_col_idx = min_col + i
            # Store both original case and lowercase for flexible matching
            self.column_mapping[col_name] = excel_col_idx
            self.column_mapping[col_name.lower()] = excel_col_idx
        
        logger.debug(f"Built column mapping: {dict((k, v) for k, v in self.column_mapping.items() if k.islower())}")
    
    def get_autocat_rules(self) -> Optional[pd.DataFrame]:
        """Load the AutoCat rules from the 'AutoCat' worksheet."""
        logger.info("Loading AutoCat rules...")
        try:
            rules_df = pd.read_excel(self.excel_file, sheet_name='AutoCat')
            logger.info(f"Loaded {len(rules_df)} rules from 'AutoCat' sheet.")
            return rules_df
        except Exception as e:
            logger.warning(f"Could not load 'AutoCat' worksheet: {e}")
            logger.debug(f"Could not find 'AutoCat' sheet. This is not an error if you don't use this feature.")
            return None

    def update_cell(self, df_index: int, column_name: str, value: Any):
        """
        Update a specific cell in the Excel table.
        
        Args:
            df_index (int): The 0-based index of the row in the DataFrame.
            column_name (str): The name of the column to update.
            value (Any): The new value to set.
        """
        try:
            if column_name not in self.column_mapping:
                logger.warning(f"Column '{column_name}' not found in the Excel table. Cannot update cell.")
                return
            
            # Excel is 1-based, header is at self.header_row, data starts below it.
            # The DataFrame index (df_index) is 0-based relative to the data.
            excel_row = self.header_row + 1 + df_index
            excel_col = self.column_mapping[column_name]
            
            logger.debug(f"Updating cell at (row={excel_row}, col={excel_col}) to '{value}'")
            
            cell = self.ws.cell(row=excel_row, column=excel_col)
            cell.value = value
            
            # Also update the in-memory dataframe to keep it consistent
            self.existing_df.loc[df_index, column_name] = value

        except Exception as e:
            logger.error(f"Failed to update cell at index {df_index}, column '{column_name}'")
            logger.debug(traceback.format_exc())

    def update_transactions(self, transactions: List[Dict[str, Any]]) -> int:
        """Add new transactions to the Transactions table."""
        if not transactions:
            logger.warning("No transactions provided")
            return 0
        
        # Check for duplicates using existing DuplicateChecker
        filtered_transactions = DuplicateChecker.check_for_duplicates(
            self.existing_df, transactions
        )
        
        if not filtered_transactions:
            logger.warning("No new transactions to import (all appear to be duplicates)")
            return 0
        
        logger.info(f"Adding {len(filtered_transactions)} new transactions...")
        
        # Validate and prepare transactions
        prepared_transactions = self._prepare_transactions(filtered_transactions)
        
        # Add transactions to the table using proper method
        rows_added = self._add_transactions_to_table(prepared_transactions)
        
        logger.info(f"Successfully added {rows_added} transactions to the table")
        return rows_added
    
    def _prepare_transactions(self, transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare transactions by validating columns and converting data types."""
        prepared = []
        
        for transaction in transactions:
            prepared_transaction = {}
            
            # Process each field in the transaction
            for key, value in transaction.items():
                # Find matching column in Excel table (case-insensitive)
                excel_col_key = self._find_matching_column(key)
                if excel_col_key is None:
                    logger.warning(f"Transaction column '{key}' not found in Excel table - skipping")
                    continue
                
                # Convert value to appropriate type
                converted_value = self._convert_value_for_excel(key, value)
                prepared_transaction[excel_col_key] = converted_value
            
            if prepared_transaction:  # Only add if we have valid columns
                prepared.append(prepared_transaction)
        
        logger.debug(f"Prepared {len(prepared)} transactions for insertion")
        return prepared
    
    def _find_matching_column(self, transaction_column: str) -> Optional[str]:
        """Find matching column in Excel table (case-insensitive)."""
        # Try exact match first
        if transaction_column in self.column_mapping:
            return transaction_column
        
        # Try case-insensitive match
        lower_key = transaction_column.lower()
        if lower_key in self.column_mapping:
            # Find the original case column name
            for original_col in self.existing_columns:
                if original_col.lower() == lower_key:
                    return original_col
        
        return None
    
    def _convert_value_for_excel(self, column_name: str, value: Any) -> Any:
        """Convert value to appropriate type for Excel storage."""
        if pd.isna(value) or value is None or value == '':
            return None
        
        # Handle date columns - ensure they're stored as datetime objects
        if self._is_date_column(column_name):
            return self._parse_date_value(value)
        
        # Handle numeric columns
        if isinstance(value, str) and value.replace('.', '').replace('-', '').replace('+', '').isdigit():
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass
        
        # Return as-is for other types
        if isinstance(value, (int, float, datetime, pd.Timestamp)):
            return value
        
        return str(value) if value is not None else None
    
    def _is_date_column(self, column_name: str) -> bool:
        """Check if column should be treated as a date column."""
        date_indicators = ['date', 'time', 'created', 'updated', 'posted']
        column_lower = column_name.lower()
        return any(indicator in column_lower for indicator in date_indicators)
    
    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        """Parse various date formats into datetime object."""
        if isinstance(value, datetime):
            return value
        elif isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        elif isinstance(value, str) and value.strip():
            try:
                # Try pandas to_datetime which handles many formats
                parsed = pd.to_datetime(value)
                return parsed.to_pydatetime() if isinstance(parsed, pd.Timestamp) else parsed
            except Exception as e:
                logger.warning(f"Could not parse date '{value}': {e}")
                return None
        
        return None
    
    def _add_transactions_to_table(self, transactions: List[Dict[str, Any]]) -> int:
        """
        Add transactions in chronological order (newest to oldest).
        Preserves formatting and table features while inserting transactions
        in the correct position based on their date.
        """
        if not transactions:
            return 0
        
        # Get current table boundaries
        min_col, min_row, max_col, current_max_row = range_boundaries(self.table.ref)
        
        # Sort transactions by date (newest first for proper insertion order)
        sorted_transactions = self._sort_transactions_by_date(transactions)
        
        # Find date column index
        date_col_idx = self._find_date_column_index(min_col)
        
        if date_col_idx is None:
            logger.warning("Date column not found, inserting transactions at the beginning")
            return self._add_transactions_at_beginning(sorted_transactions, min_col, min_row, max_col, current_max_row)
        
        # Insert transactions one by one in the correct chronological position
        rows_inserted = 0
        for transaction in sorted_transactions:
            insert_row = self._find_insertion_position(transaction, date_col_idx, min_row, current_max_row + rows_inserted)
            
            # Insert new row
            self.ws.insert_rows(insert_row)
            rows_inserted += 1
            
            # Copy formatting from adjacent row
            self._copy_row_formatting(insert_row - 1 if insert_row > min_row + 1 else insert_row + 1, 
                                    insert_row, min_col, max_col)
            
            # Populate the row with data
            self._populate_table_row(insert_row, transaction, min_col)
            
            # Log for debugging
            self._log_transaction_insertion(transaction, insert_row)
        
        # Update table range to include new rows
        new_max_row = current_max_row + rows_inserted
        new_table_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
        self.table.ref = new_table_range
        logger.debug(f"Updated table range to: {new_table_range}")
        
        # Extend formatting and validation rules
        self._extend_conditional_formatting(min_col, min_row, max_col, current_max_row, new_max_row)
        self._extend_data_validation(min_col, min_row, max_col, current_max_row, new_max_row)
        
        return rows_inserted
    
    def _find_date_column_index(self, min_col: int) -> Optional[int]:
        """Find the index of the date column in the table."""
        for i, col_name in enumerate(self.existing_columns):
            if self._is_date_column(col_name):
                return min_col + i
        return None
    
    def _find_insertion_position(self, transaction: Dict[str, Any], date_col_idx: int, 
                               header_row: int, max_row: int) -> int:
        """
        Find the correct insertion position for a transaction based on date.
        Maintains newest-to-oldest order.
        """
        trans_date = self._extract_transaction_date(transaction)
        if trans_date is None:
            # If no date, insert at the beginning after header
            return header_row + 1
        
        # Find insertion point by comparing with existing dates
        for row in range(header_row + 1, max_row + 1):
            cell_value = self.ws.cell(row=row, column=date_col_idx).value
            existing_date = self._parse_excel_date(cell_value)
            
            if existing_date is None:
                # Empty row or unparseable date, insert here
                return row
            
            # If transaction date is newer than existing date, insert here
            if trans_date > existing_date:
                return row
        
        # If we get here, insert at the end
        return max_row + 1
    
    def _extract_transaction_date(self, transaction: Dict[str, Any]) -> Optional[datetime]:
        """Extract date from transaction data."""
        for key, value in transaction.items():
            if self._is_date_column(key):
                if isinstance(value, datetime):
                    return value
                elif value is not None:
                    return self._parse_date_value(value)
        return None
    
    def _parse_excel_date(self, cell_value: Any) -> Optional[datetime]:
        """Parse date value from Excel cell."""
        if cell_value is None:
            return None
        
        if isinstance(cell_value, datetime):
            return cell_value
        elif isinstance(cell_value, (int, float)):
            # Excel date serial number
            try:
                return pd.to_datetime(cell_value, origin='1899-12-30', unit='D')
            except:
                return None
        elif isinstance(cell_value, str):
            return self._parse_date_value(cell_value)
        
        return None
    
    def _add_transactions_at_beginning(self, transactions: List[Dict[str, Any]], 
                                     min_col: int, min_row: int, max_col: int, current_max_row: int) -> int:
        """
        Fallback method to add all transactions at the beginning of the table.
        Used when date column is not found.
        """
        insert_position = min_row + 1  # Right after header
        
        for i, transaction in enumerate(transactions):
            # Insert row at the beginning
            self.ws.insert_rows(insert_position)
            
            # Copy formatting
            source_row = insert_position + 1 if insert_position + 1 <= current_max_row + i else current_max_row
            self._copy_row_formatting(source_row, insert_position, min_col, max_col)
            
            # Populate row
            self._populate_table_row(insert_position, transaction, min_col)
            
            # Log transaction
            self._log_transaction_insertion(transaction, insert_position)
        
        # Update table range
        new_max_row = current_max_row + len(transactions)
        new_table_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
        self.table.ref = new_table_range
        
        # Extend formatting
        self._extend_conditional_formatting(min_col, min_row, max_col, current_max_row, new_max_row)
        self._extend_data_validation(min_col, min_row, max_col, current_max_row, new_max_row)
        
        return len(transactions)
    
    def _log_transaction_insertion(self, transaction: Dict[str, Any], row_num: int):
        """Log transaction insertion for debugging."""
        date_val = 'No Date'
        amount_val = 'No Amount'
        desc_val = 'No Description'
        
        for key, value in transaction.items():
            if self._is_date_column(key):
                date_val = str(value) if value else 'No Date'
            elif 'amount' in key.lower():
                amount_val = str(value) if value else 'No Amount'
            elif 'description' in key.lower() or 'desc' in key.lower():
                desc_val = str(value)[:50] if value else 'No Description'
        
        logger.debug(f"  Inserted at row {row_num}: {date_val} | {amount_val} | {desc_val}")
    
    def _copy_row_formatting(self, source_row: int, target_row: int, min_col: int, max_col: int):
        """Copy formatting from source row to target row."""
        for col in range(min_col, max_col + 1):
            try:
                source_cell = self.ws.cell(row=source_row, column=col)
                target_cell = self.ws.cell(row=target_row, column=col)
                
                # Copy cell formatting
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
            except Exception as e:
                logger.warning(f"Failed to copy formatting from row {source_row} to {target_row}, col {col}: {e}")
    
    def _extend_conditional_formatting(self, min_col: int, min_row: int, max_col: int, old_max_row: int, new_max_row: int):
        """Extend conditional formatting rules that apply to entire table columns."""
        first_data_row = min_row + 1
        rules_updated = 0
        rules_to_recreate = []
        cf_rules_to_remove = []
        
        # Process existing conditional formatting rules
        for cf in list(self.ws.conditional_formatting):
            for range_obj in list(cf.cells.ranges):
                try:
                    range_str = range_obj.coord
                    cf_min_col, cf_min_row, cf_max_col, cf_max_row = range_boundaries(range_str)
                    
                    # Check if this is a complete column rule that needs extending
                    if (cf_min_col == cf_max_col and  # Single column
                        cf_min_col >= min_col and cf_max_col <= max_col and  # Within table
                        cf_min_row == first_data_row and  # Starts at first data row
                        cf_max_row >= old_max_row - 5):  # Ends near old table end (some tolerance)
                        
                        new_range_str = f"{get_column_letter(cf_min_col)}{first_data_row}:{get_column_letter(cf_max_col)}{new_max_row}"
                        
                        for rule in cf.rules:
                            rules_to_recreate.append((new_range_str, rule))
                            rules_updated += 1
                        
                        # Mark for removal
                        cf_rules_to_remove.append(cf)
                        logger.debug(f"Extended conditional formatting from {range_str} to {new_range_str}")
                        break
                        
                except Exception as e:
                    logger.error(f"Failed to process conditional formatting: {e}")
                    continue
        
        # Remove old rules using the correct method
        for cf_rule in cf_rules_to_remove:
            try:
                # Use the internal _cf_rules dictionary to remove rules
                if hasattr(self.ws.conditional_formatting, '_cf_rules'):
                    if cf_rule in self.ws.conditional_formatting._cf_rules:
                        del self.ws.conditional_formatting._cf_rules[cf_rule]
                else:
                    # Fallback: try to clear and rebuild all rules
                    logger.warning("Could not remove specific conditional formatting rule, skipping removal")
            except Exception as e:
                logger.warning(f"Could not remove conditional formatting rule: {e}")
        
        # Recreate rules with extended ranges
        for range_str, rule in rules_to_recreate:
            try:
                self.ws.conditional_formatting.add(range_str, rule)
            except Exception as e:
                logger.error(f"Failed to recreate conditional formatting rule: {e}")
        
        if rules_updated > 0:
            logger.debug(f"Extended {rules_updated} conditional formatting rules")
    
    def _extend_data_validation(self, min_col: int, min_row: int, max_col: int, old_max_row: int, new_max_row: int):
        """Extend data validation rules that apply to entire table columns."""
        first_data_row = min_row + 1
        rules_updated = 0
        validations_to_recreate = []
        
        # Find data validation rules that need updating
        for dv in list(self.ws.data_validations.dataValidation):
            for range_obj in list(dv.cells.ranges):
                try:
                    range_str = range_obj.coord
                    dv_min_col, dv_min_row, dv_max_col, dv_max_row = range_boundaries(range_str)
                    
                    if (dv_min_col == dv_max_col and  # Single column
                        dv_min_col >= min_col and dv_max_col <= max_col and  # Within table
                        dv_min_row == first_data_row and  # Starts at first data row
                        dv_max_row >= old_max_row - 5):  # Ends near old table end
                        
                        new_range_str = f"{get_column_letter(dv_min_col)}{first_data_row}:{get_column_letter(dv_max_col)}{new_max_row}"
                        validations_to_recreate.append((dv, new_range_str))
                        rules_updated += 1
                        logger.debug(f"Will recreate data validation for range {new_range_str}")
                        break
                        
                except Exception as e:
                    logger.error(f"Error processing data validation: {e}")
                    continue
        
        # Remove old validations and add new ones
        for old_dv, new_range_str in validations_to_recreate:
            try:
                self.ws.data_validations.dataValidation.remove(old_dv)
                
                # Create new validation with same properties
                new_dv = DataValidation(
                    type=old_dv.type,
                    formula1=old_dv.formula1,
                    formula2=old_dv.formula2,
                    showErrorMessage=old_dv.showErrorMessage,
                    errorTitle=old_dv.errorTitle,
                    error=old_dv.error,
                    showInputMessage=old_dv.showInputMessage,
                    promptTitle=old_dv.promptTitle,
                    prompt=old_dv.prompt,
                    allow_blank=old_dv.allow_blank
                )
                
                new_dv.add(new_range_str)
                self.ws.data_validations.append(new_dv)
                
            except Exception as e:
                logger.error(f"Failed to recreate data validation: {e}")
        
        if rules_updated > 0:
            logger.debug(f"Extended {rules_updated} data validation rules")
    
    def _populate_table_row(self, row_num: int, transaction: Dict[str, Any], start_col: int):
        """Populate a table row with transaction data, handling extra columns gracefully."""
        for i, col_name in enumerate(self.existing_columns):
            col_idx = start_col + i
            cell = self.ws.cell(row=row_num, column=col_idx)
            
            # Get value from transaction (case-insensitive lookup)
            value = transaction.get(col_name)
            if value is None:
                # Try case-insensitive lookup
                for trans_key, trans_value in transaction.items():
                    if trans_key.lower() == col_name.lower():
                        value = trans_value
                        break
            
            # Set cell value
            if pd.isna(value) or value is None:
                cell.value = None
            elif isinstance(value, (pd.Timestamp, datetime)):
                cell.value = value
                # Ensure date formatting is preserved
                if cell.number_format == 'General':
                    cell.number_format = 'mm/dd/yyyy'
            elif isinstance(value, (int, float)):
                cell.value = value
            else:
                cell.value = str(value) if value is not None else None
    
    def _sort_transactions_by_date(self, transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Sort transactions by date, newest first.
        This ensures proper insertion order for chronological placement.
        """
        def get_date_key(transaction):
            trans_date = self._extract_transaction_date(transaction)
            return trans_date if trans_date else datetime.min
        
        sorted_txns = sorted(transactions, key=get_date_key, reverse=True)
        logger.debug(f"Sorted {len(sorted_txns)} transactions by date (newest first)")
        return sorted_txns
    
    def save(self):
        """Save the workbook."""
        logger.info(f"Saving workbook to: {self.excel_file}")
        try:
            self.wb.save(self.excel_file)
            logger.info("Workbook saved successfully")
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
        finally:
            if self.wb:
                self.wb.close()
                logger.info("Workbook closed")