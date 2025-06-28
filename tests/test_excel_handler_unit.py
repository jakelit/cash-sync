import pytest
from excel_finance_tools.excel_handler import ExcelHandler
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import tempfile

class TestExcelHandlerUnit:
    """Unit tests for ExcelHandler class."""

    @pytest.mark.unit
    def test_load_workbook_valid(self, tmp_path):
        """UT036: Valid Excel file - Should load workbook successfully."""        
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2024-01-01", 10.0, "Test"])
        tab = Table(displayName="Transactions", ref="A1:C2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(str(excel_file))
        handler = ExcelHandler(str(excel_file))
        handler.load_workbook()
        assert handler.wb is not None
        assert handler.ws is not None

    @pytest.mark.unit
    def test_load_workbook_missing_file(self, tmp_path):
        """UT037: Missing Excel file - Should raise FileNotFoundError."""
        missing_file = tmp_path / "no_such.xlsx"
        handler = ExcelHandler(str(missing_file))
        with pytest.raises(FileNotFoundError):
            handler.load_workbook()

    @pytest.mark.unit
    def test_load_workbook_missing_table(self, tmp_path):
        """UT038: Missing Transactions table - Should raise ValueError if table is missing."""
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NotTransactions"
        ws.append(["Date", "Amount", "Description"])
        wb.save(str(excel_file))
        handler = ExcelHandler(str(excel_file))
        with pytest.raises(ValueError):
            handler.load_workbook()

    @pytest.mark.unit
    def test_update_transactions_valid(self):
        """UT039: Valid transaction list - Should add transactions successfully."""
        # Create a dummy Excel file with Transactions table
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            excel_file = tmp.name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2024-01-01", 10.0, "Test"])
        tab = Table(displayName="Transactions", ref="A1:C2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(excel_file)
        handler = ExcelHandler(excel_file)
        handler.load_workbook()
        handler.existing_columns = ["Date", "Amount", "Description"]
        handler.existing_df = pd.DataFrame([])
        txns = [{"Date": "2024-01-01", "Amount": 10.0, "Description": "Test"}]
        count = handler.update_transactions(txns)
        assert count == 1

    @pytest.mark.unit
    def test_update_transactions_empty(self):
        """UT040: Empty transaction list - Should not add any transactions."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            excel_file = tmp.name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2024-01-01", 10.0, "Test"])
        tab = Table(displayName="Transactions", ref="A1:C2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(excel_file)
        handler = ExcelHandler(excel_file)
        handler.load_workbook()
        handler.existing_columns = ["Date", "Amount", "Description"]
        handler.existing_df = pd.DataFrame([])
        txns = []
        count = handler.update_transactions(txns)
        assert count == 0

    @pytest.mark.unit
    def test_update_transactions_all_duplicates(self, tmp_path):
        """UT041: All duplicates - Should return 0 when all transactions are duplicates of existing ones."""
        # Create a dummy Excel file with one transaction in the Transactions table
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description", "Full Description"])
        ws.append(["2024-01-01", 10.0, "Test", "Test Full Desc"])
        tab = Table(displayName="Transactions", ref="A1:D2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(str(excel_file))
        handler = ExcelHandler(str(excel_file))
        handler.load_workbook()
        handler.existing_columns = ["Date", "Amount", "Description", "Full Description"]
        handler.existing_df = pd.DataFrame([
            {"Date": "2024-01-01", "Amount": 10.0, "Description": "Test", "Full Description": "Test Full Desc"}
        ])
        txns = [{"Date": "2024-01-01", "Amount": 10.0, "Description": "Test", "Full Description": "Test Full Desc"}]
        count = handler.update_transactions(txns)
        assert count == 0 

    @pytest.mark.unit
    def test_prepare_transactions_valid(self):
        """UT042: Valid transaction preparation - Should prepare transaction data with correct types and columns."""
        handler = ExcelHandler("dummy.xlsx")
        handler.existing_columns = ["Date", "Amount", "Description"]
        handler.column_mapping = {"Date": 1, "Amount": 2, "Description": 3, "date": 1, "amount": 2, "description": 3}
        txns = [{"Date": "2024-01-01", "Amount": "10.00", "Description": "Test"}]
        prepared = handler._prepare_transactions(txns)
        assert len(prepared) == 1
        assert set(prepared[0].keys()) == {"Date", "Amount", "Description"}
        assert isinstance(prepared[0]["Amount"], float)

    @pytest.mark.unit
    def test_prepare_transactions_invalid_column(self):
        """UT043: Invalid column mapping - Should skip unmapped columns."""
        handler = ExcelHandler("dummy.xlsx")
        handler.existing_columns = ["Date", "Amount"]
        handler.column_mapping = {"Date": 1, "Amount": 2, "date": 1, "amount": 2}
        txns = [{"Date": "2024-01-01", "Amount": "10.00", "Unknown": "foo"}]
        prepared = handler._prepare_transactions(txns)
        assert len(prepared) == 1
        assert set(prepared[0].keys()) == {"Date", "Amount"}
        assert "Unknown" not in prepared[0]

    @pytest.mark.unit
    def test_find_matching_column_exact(self):
        """UT044: Exact column match - Should return exact column name if present."""
        handler = ExcelHandler("dummy.xlsx")
        handler.existing_columns = ["Date", "Amount"]
        handler.column_mapping = {"Date": 1, "Amount": 2, "date": 1, "amount": 2}
        assert handler._find_matching_column("Date") == "Date"
        assert handler._find_matching_column("Amount") == "Amount"

    @pytest.mark.unit
    def test_find_matching_column_case_insensitive(self):
        """UT045: Case-insensitive match - Should return correct case column name."""
        handler = ExcelHandler("dummy.xlsx")
        handler.existing_columns = ["Date", "Amount"]
        handler.column_mapping = {"Date": 1, "Amount": 2, "date": 1, "amount": 2}
        assert handler._find_matching_column("date") == "date"
        assert handler._find_matching_column("amount") == "amount"

    @pytest.mark.unit
    def test_convert_value_for_excel_date(self):
        """UT046: Date value conversion - Should convert date string to datetime object."""
        handler = ExcelHandler("dummy.xlsx")
        handler._is_date_column = lambda col: True  # Force date column
        dt = handler._convert_value_for_excel("Date", "2024-01-01")
        from datetime import datetime
        assert isinstance(dt, datetime)
        assert dt.year == 2024 and dt.month == 1 and dt.day == 1

    @pytest.mark.unit
    def test_convert_value_for_excel_numeric(self):
        """UT047: Numeric value conversion - Should convert numeric string to float or int."""
        handler = ExcelHandler("dummy.xlsx")
        handler._is_date_column = lambda col: False
        assert handler._convert_value_for_excel("Amount", "123.45") == 123.45
        assert handler._convert_value_for_excel("Amount", "42") == 42

    @pytest.mark.unit
    def test_convert_value_for_excel_string(self):
        """UT048: String value conversion - Should leave string value unchanged if not date or numeric."""
        handler = ExcelHandler("dummy.xlsx")
        handler._is_date_column = lambda col: False
        assert handler._convert_value_for_excel("Description", "Test") == "Test"

    @pytest.mark.unit
    def test_is_date_column_true(self):
        """UT049: Date column detection - Should return True for columns with date indicators."""
        handler = ExcelHandler("dummy.xlsx")
        assert handler._is_date_column("Date")
        assert handler._is_date_column("Transaction Time")
        assert handler._is_date_column("Created At")
        assert handler._is_date_column("Updated On")
        assert handler._is_date_column("Posted Date")

    @pytest.mark.unit
    def test_is_date_column_false(self):
        """UT050: Non-date column detection - Should return False for columns without date indicators."""
        handler = ExcelHandler("dummy.xlsx")
        assert not handler._is_date_column("Amount")
        assert not handler._is_date_column("Description")
        assert not handler._is_date_column("Category")

    @pytest.mark.unit
    @pytest.mark.parametrize("exception_type", [
        ValueError, FileNotFoundError, OSError, pd.errors.EmptyDataError, pd.errors.ParserError
    ])
    def test_get_autocat_rules_error_handling(self, mocker, tmp_path, exception_type):
        """UT073: get_autocat_rules() handles missing/invalid AutoCat worksheet and returns None on error."""
        # Create a dummy Excel file
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2024-01-01", 10.0, "Test"])
        tab = Table(displayName="Transactions", ref="A1:C2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(str(excel_file))
        handler = ExcelHandler(str(excel_file))
        # Patch pd.read_excel to raise the exception
        mocker.patch("pandas.read_excel", side_effect=exception_type)
        result = handler.get_autocat_rules()
        assert result is None

    @pytest.mark.unit
    def test_update_cell_missing_column(self, mocker, tmp_path):
        """UT074: update_cell() handles missing column by logging warning and returning early."""
        # Create a dummy Excel file
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2024-01-01", 10.0, "Test"])
        tab = Table(displayName="Transactions", ref="A1:C2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(str(excel_file))
        handler = ExcelHandler(str(excel_file))
        handler.load_workbook()
        handler.column_mapping = {"Date": 1, "Amount": 2, "Description": 3}  # No 'Category'
        handler.existing_df = mocker.Mock()
        handler.ws = mocker.Mock()
        # Patch logger to check for warning
        mock_logger = mocker.patch("excel_finance_tools.excel_handler.logger")
        # Call update_cell with a missing column
        result = handler.update_cell(0, "Category", "TestValue")
        # Should log a warning and return early
        mock_logger.warning.assert_called_once_with(
            "Column '%s' not found in the Excel table. Cannot update cell.", "Category"
        )
        # Should not attempt to update worksheet
        handler.ws.cell.assert_not_called()
        assert result is None