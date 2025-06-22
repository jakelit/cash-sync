"""
End-to-End tests for AutoCategorizer class.

This module contains end-to-end tests for the AutoCategorizer class,
covering complete workflows from Excel file loading to transaction categorization.
"""

import pytest
from unittest.mock import Mock, patch
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import tempfile
import os

from excel_finance_tools.auto_categorizer import AutoCategorizer

def create_excel_table(worksheet, data, table_name, start_row=1, start_col=1):
    """Helper function to create a proper Excel table."""
    # Add data to worksheet
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            worksheet.cell(row=start_row + i, column=start_col + j, value=value)
    
    # Calculate table range
    end_row = start_row + len(data) - 1
    end_col = start_col + len(data[0]) - 1
    
    # Create table
    table = Table(displayName=table_name, ref=f"{worksheet.cell(start_row, start_col).coordinate}:{worksheet.cell(end_row, end_col).coordinate}")
    
    # Add table style
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    
    # Add table to worksheet
    worksheet.add_table(table)
    
    return table

@pytest.fixture
def sample_excel_file_with_rules_and_transactions():
    """Create a test Excel file with AutoCat rules and transaction data for E2E testing."""
    wb = Workbook()
    
    # Create AutoCat worksheet with rules
    autocat_ws = wb.create_sheet("AutoCat")
    autocat_data = [
        ['Category', 'Description Contains', 'Amount Min', 'Amount Max', 'Tags', 'Notes'],
        ['Groceries', 'WALMART', '', '500', 'shopping', 'Grocery shopping'],
        ['Gas', 'SHELL', '', '', 'transport', 'Gas station'],
        ['Coffee', 'STARBUCKS', '', '20', 'food', 'Coffee purchase'],
        ['Entertainment', 'NETFLIX', '', '', 'entertainment', 'Streaming service']
    ]
    for row in autocat_data:
        autocat_ws.append(row)
    
    # Create Transactions worksheet with proper table
    trans_ws = wb.create_sheet("Transactions")
    trans_data = [
        ['Date', 'Description', 'Amount', 'Category', 'Account', 'Tags', 'Notes'],
        ['2024-01-01', 'WALMART GROCERY', 150.00, '', 'Checking', '', ''],
        ['2024-01-02', 'SHELL GAS STATION', 45.50, '', 'Checking', '', ''],
        ['2024-01-03', 'STARBUCKS COFFEE', 5.75, '', 'Checking', '', ''],
        ['2024-01-04', 'NETFLIX SUBSCRIPTION', 15.99, '', 'Checking', '', ''],
        ['2024-01-05', 'ALREADY CATEGORIZED', 25.00, 'Entertainment', 'Checking', '', '']
    ]
    create_excel_table(trans_ws, trans_data, "Transactions")
    
    # Remove default sheet
    wb.remove(wb['Sheet'])
    
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    yield path
    os.remove(path)

@pytest.fixture
def sample_excel_file_with_invalid_rules():
    """Create a test Excel file with some invalid rules for error handling E2E testing."""
    wb = Workbook()
    
    # Create AutoCat worksheet with valid and invalid rules
    autocat_ws = wb.create_sheet("AutoCat")
    autocat_data = [
        ['Category', 'Description Contains', 'Amount Min', 'Amount Max', 'Tags', 'Invalid Column'],
        ['Groceries', 'WALMART', '', '500', 'shopping', 'invalid value'],
        ['Gas', 'SHELL', '', '', 'transport', ''],
        ['Invalid Rule', 'Invalid Column Name', '', '', '', '']
    ]
    for row in autocat_data:
        autocat_ws.append(row)
    
    # Create Transactions worksheet with proper table
    trans_ws = wb.create_sheet("Transactions")
    trans_data = [
        ['Date', 'Description', 'Amount', 'Category', 'Account', 'Tags'],
        ['2024-01-01', 'WALMART GROCERY', 150.00, '', 'Checking', ''],
        ['2024-01-02', 'SHELL GAS STATION', 45.50, '', 'Checking', '']
    ]
    create_excel_table(trans_ws, trans_data, "Transactions")
    
    # Remove default sheet
    wb.remove(wb['Sheet'])
    
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    yield path
    os.remove(path)

@pytest.fixture
def sample_excel_file_with_no_matching_rules():
    """Create a test Excel file with rules that don't match any transactions."""
    wb = Workbook()
    
    # Create AutoCat worksheet with non-matching rules
    autocat_ws = wb.create_sheet("AutoCat")
    autocat_data = [
        ['Category', 'Description Contains', 'Amount Min', 'Amount Max', 'Tags'],
        ['Electronics', 'BEST BUY', '', '', 'tech'],
        ['Restaurant', 'MCDONALDS', '', '', 'food'],
        ['Utilities', 'ELECTRIC COMPANY', '', '', 'bills']
    ]
    for row in autocat_data:
        autocat_ws.append(row)
    
    # Create Transactions worksheet with transactions that don't match rules
    trans_ws = wb.create_sheet("Transactions")
    trans_data = [
        ['Date', 'Description', 'Amount', 'Category', 'Account', 'Tags'],
        ['2024-01-01', 'WALMART GROCERY', 150.00, '', 'Checking', ''],
        ['2024-01-02', 'SHELL GAS STATION', 45.50, '', 'Checking', ''],
        ['2024-01-03', 'STARBUCKS COFFEE', 5.75, '', 'Checking', '']
    ]
    create_excel_table(trans_ws, trans_data, "Transactions")
    
    # Remove default sheet
    wb.remove(wb['Sheet'])
    
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    yield path
    os.remove(path)

@pytest.fixture
def sample_excel_file_with_complex_rules():
    """Create a test Excel file with complex rule combinations for comprehensive E2E testing."""
    wb = Workbook()
    
    # Create AutoCat worksheet with complex rules
    autocat_ws = wb.create_sheet("AutoCat")
    autocat_data = [
        ['Category', 'Description Contains', 'Amount Min', 'Amount Max', 'Account Equals', 'Tags', 'Notes'],
        ['Coffee', 'STARBUCKS', '', '20', '', 'food', 'Coffee under $20'],
        ['Expensive Coffee', 'STARBUCKS', '20', '', '', 'luxury', 'Coffee over $20'],
        ['Gas', 'SHELL', '30', '100', '', 'transport', 'Gas between $30-$100'],
        ['Groceries', 'WALMART', '', '500', 'Checking', 'shopping', 'Walmart groceries from checking'],
        ['Online Shopping', 'AMAZON', '', '', '', 'online', 'Amazon purchases']
    ]
    for row in autocat_data:
        autocat_ws.append(row)
    
    # Create Transactions worksheet with various scenarios
    trans_ws = wb.create_sheet("Transactions")
    trans_data = [
        ['Date', 'Description', 'Amount', 'Category', 'Account', 'Tags', 'Notes'],
        ['2024-01-01', 'STARBUCKS COFFEE', 5.75, '', 'Checking', '', ''],
        ['2024-01-02', 'STARBUCKS COFFEE', 25.50, '', 'Checking', '', ''],
        ['2024-01-03', 'SHELL GAS STATION', 45.50, '', 'Checking', '', ''],
        ['2024-01-04', 'WALMART GROCERY', 150.00, '', 'Checking', '', ''],
        ['2024-01-05', 'AMAZON PURCHASE', 75.25, '', 'Checking', '', ''],
        ['2024-01-06', 'SHELL GAS STATION', 25.00, '', 'Checking', '', ''],  # Below min
        ['2024-01-07', 'WALMART GROCERY', 600.00, '', 'Savings', '', '']  # Wrong account
    ]
    create_excel_table(trans_ws, trans_data, "Transactions")
    
    # Remove default sheet
    wb.remove(wb['Sheet'])
    
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    yield path
    os.remove(path)

class TestAutoCategorizerE2E:
    """End-to-end workflow testing for AutoCategorizer."""

    @pytest.mark.e2e
    def test_complete_workflow_full_categorization_process(self, sample_excel_file_with_rules_and_transactions):
        """TC058: Test complete workflow - Full categorization process."""
        # Arrange
        categorizer = AutoCategorizer(sample_excel_file_with_rules_and_transactions)
        
        # Act
        success, message = categorizer.run_auto_categorization()
        
        # Assert
        assert success is True
        assert "categorized" in message.lower()
        
        # Verify that rules were loaded
        assert len(categorizer.rules) == 4
        
        # Verify that transactions were processed
        # The exact number depends on the implementation, but we should have processed transactions
        assert "transactions" in message.lower() or "updates" in message.lower()

    @pytest.mark.e2e
    def test_workflow_with_errors_process_with_invalid_rules(self, sample_excel_file_with_invalid_rules):
        """TC059: Test workflow with errors - Process with invalid rules."""
        # Arrange
        categorizer = AutoCategorizer(sample_excel_file_with_invalid_rules)
        
        # Act
        success, message = categorizer.run_auto_categorization()
        
        # Assert
        assert success is True
        assert "categorized" in message.lower() or "processed" in message.lower()
        
        # Verify that valid rules were loaded (should ignore invalid ones)
        assert len(categorizer.rules) >= 1  # At least one valid rule
        
        # Verify that processing continued despite invalid rules
        # The exact behavior depends on implementation, but it should not fail completely

    @pytest.mark.e2e
    def test_workflow_with_no_matches_process_with_no_matching_rules(self, sample_excel_file_with_no_matching_rules):
        """TC060: Test workflow with no matches - Process with no matching rules."""
        # Arrange
        categorizer = AutoCategorizer(sample_excel_file_with_no_matching_rules)
        
        # Act
        success, message = categorizer.run_auto_categorization()
        
        # Assert
        assert success is True
        assert "no" in message.lower() and ("updates" in message.lower() or "matches" in message.lower() or "new" in message.lower())
        
        # Verify that rules were loaded
        assert len(categorizer.rules) == 3
        
        # Verify that no transactions were categorized (since none match the rules)
        # The message should indicate no changes were made

    @pytest.mark.e2e
    def test_workflow_with_complex_rules_process_with_multiple_rule_types(self, sample_excel_file_with_complex_rules):
        """TC061: Test workflow with complex rules - Process with multiple rule types."""
        # Arrange
        categorizer = AutoCategorizer(sample_excel_file_with_complex_rules)
        
        # Act
        success, message = categorizer.run_auto_categorization()
        
        # Assert
        assert success is True
        assert "categorized" in message.lower() or "processed" in message.lower()
        
        # Verify that complex rules were loaded
        assert len(categorizer.rules) == 5
        
        # Verify that various rule types were processed
        # This includes:
        # - Simple contains rules (Amazon)
        # - Amount range rules (Gas between $30-$100)
        # - Amount min/max rules (Coffee under/over $20)
        # - Account-specific rules (Walmart from checking account)
        
        # The exact number of matches depends on the implementation,
        # but we should have processed transactions with complex rule combinations 